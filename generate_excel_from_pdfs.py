import os  
import sys  
from openai import AzureOpenAI  
from pdfminer.high_level import extract_text  
import openpyxl  
from openpyxl.worksheet.datavalidation import DataValidation  

# Function to call Azure OpenAI GPT-4 for extracting questions and answers  
def extract_questions_and_answers_gpt(text):  
    deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")  
    client = AzureOpenAI(  
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),  
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),  
        api_version="2024-02-01"  
    )  
      
    prompt = f"""  
You are an expert at extracting structured data from documents. Please identify and extract questions and their possible answers from the following text. All questions have three numbers in front of them (e.g., 4.1.3.). Each question can be open-ended, a Yes-No question, or a multiple-choice question.   
  
- For open-ended questions, the text "Open-ended" will appear below the question.  
- For Yes-No questions, the answers will be "Yes" and "No".  
- For multiple-choice questions, the answers will be bullet points.  
  
Please format the output as follows:  
Section Title | Question Number | Question | Possible Answers (comma-separated) | Scoring  
  
Ensure:  
- Sections are clearly marked.  
- The output contains no additional text or explanations.  
- Each question and its possible answers are formatted exactly as specified.  
  
Example output:  
Section Title | 3.1.1. | What is the capital of France? | Paris, London, Berlin | Paris - Correct, London - Incorrect, Berlin - Incorrect  
Section Title | 4.2. | Is the sky blue? | Yes, No | Yes - Correct, No - Incorrect  
  
Extract from the following text:  
{text}  
"""  
  
    response = client.chat.completions.create(  
        model=deployment,  
        messages=[  
            {"role": "system", "content": "Assistant is a large language model trained by OpenAI."},  
            {"role": "user", "content": prompt}  
        ]  
    )  
      
    # Extract content from response  
    gpt_response = response.choices[0].message.content.strip()  
    return gpt_response  
  
def extract_questions_and_answers(pdf_path):  
    print(f"Extracting text from PDF: {pdf_path}")  
    text = extract_text(pdf_path)  
    print("Text extracted, calling GPT-4...")  
    gpt_response = extract_questions_and_answers_gpt(text)  
    qa_list = []  
    lines = gpt_response.split('\n')  
    current_section = None  
  
    for line in lines:  
        parts = [part.strip() for part in line.split('|')]  
        if len(parts) == 1:  # Section Title  
            current_section = parts[0]  
            qa_list.append((current_section, 'Section Title', '', '', ''))  
        elif len(parts) == 5:  
            section_title, question_number, question, possible_answers, scoring = parts  
            qa_list.append((section_title, question_number, question, possible_answers, scoring))  
        else:  
            print(f"Skipping line due to incorrect format: {line}")  
    return qa_list  

def create_excel(qa_data, sheet_name, wb):  
    ws = wb.create_sheet(title=sheet_name)  
    ws.append(["Question Number", "The question", "Picklist", "Scoring", "Assigned To"])  
    # Set column headers to bold  
    for cell in ws[1]:  
        cell.font = openpyxl.styles.Font(bold=True)  
    for qa in qa_data:  
        section_title, question_number, question, possible_answers, scoring = qa  
        if question_number == 'Section Title':  
            ws.append([section_title, '', '', '', ''])  
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)  
        else:  
            picklist = possible_answers  
            ws.append([question_number, question, picklist, scoring, ""])  

def process_pdfs_in_directory(pdf_dir, output_file):  
    wb = openpyxl.Workbook()
    for root, _, files in os.walk(pdf_dir):  
        for pdf_file in files:  
            if pdf_file.endswith(".pdf"):  
                pdf_path = os.path.join(root, pdf_file)  
                qa_data = extract_questions_and_answers(pdf_path)  
                sheet_name = os.path.splitext(pdf_file)[0][:31]  # Sheet names must be <= 31 chars  
                create_excel(qa_data, sheet_name, wb)
    # Remove the default sheet created by openpyxl  
    if 'Sheet' in wb.sheetnames:  
        wb.remove(wb['Sheet'])
    wb.save(output_file)  
  
def main(pdf_dir):  
    output_file = os.path.join(pdf_dir, "combined_questions_answers.xlsx")  
    process_pdfs_in_directory(pdf_dir, output_file)  
  
if __name__ == "__main__":  
    try:  
        pdf_dir = sys.argv[1]  
        main(pdf_dir)  
        print("PDFs have been extracted to a combined Excel file")  
    except Exception as e:  
        print("Error:", str(e))
